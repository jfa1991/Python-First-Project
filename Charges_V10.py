from tkinter import *
from tkinter import ttk
from tkinter import messagebox
import sqlite3, requests,json,openpyxl,os,time, datetime,random
import matplotlib.pyplot as plt; plt.rcdefaults()
import matplotlib.pyplot as plt
import numpy as np
import tkinter.font as tkFont
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font


root = Tk()
root.title('Financial App')
root.geometry('550x207')

font_style = tkFont.Font(family = 'Times New Roman', size = 15, weight = 'bold')


# Create Data Base bookkeeping

conn = sqlite3.connect('/Users/Kappadona/Desktop/bookkeeping_v10.db')
cur = conn.cursor()

cur.execute(""" CREATE TABLE IF NOT EXISTS bookkeeping_v10 (
    spesa text,
    natural_del_costo text,
    costo_EUR int,
    costo_CHF int,
    daily_exchange_rate REAL,
    data text)""")


conn.commit()
conn.close()

# Create Data Base budget

conn_1 = sqlite3.connect('/Users/Kappadona/Desktop/budget_v4.db')
cur_1 = conn_1.cursor()

cur_1.execute(""" Create TABLE IF NOT EXISTS budget_v4 (
      tipo_budget text,
      natura_budget text,
      data_inizio int,
      data_fine int,
      somma_budget int)""")

conn_1.commit()
conn_1.close()

# Set up data for dropdown menu type of payment

fees = {'Prepaid/Credit Card': 0.0175, 'Cash con ritiro al bancomat': 0.04,'Pagamento Posta':0}
clicked_type_of_payment = StringVar()

# Set up data for dropdown menu to select nature of costs and
# combobox for selecting type of costs

costs_nature = {'Acquisti Personali': ['Accessori','Altro','Calzature','Giocatoli','Profumi','Vestiti'],
                'Attività Recreative e Intratenimento':['Altro','Bar','Libri','Magazine','Ristorante'],
                'Altro':['Altro',''],
                'Comunicazione e Media':['Altro','Internet','Telefono'],
                'Costi Energetici': ['Acqua', 'Altro','Gas', 'Luce', 'Riscaldamento'],
                'Salute': ['Altro','Medicinali'],'Spese Domestiche': ['Altro','Acquisti per casa'],
                'Spostamenti e Transporti':['Altro','Assicurazione Macchina','Assicurazione Vespa','Biglietti Bus', 'Biglietti Nave', 'Biglietti Treno', 'Bullo','Carburante','Costi di stazionamento','Servizi e Riparazione','Taxi'],
                'Vacanze e Viaggi': ['Escursioni e Uscite','Hotel']}

# Set up data for drowpdown menu type of budget

list_type_of_budget = [None,'Budget generale', 'Budget per natura dei costi']
clicked_budget = StringVar()


# Create function to determine whether numeric string is an integer or float
# Would be used in Function submit_data_costs

def is_integer(n):
    try:
        float(n)
    except ValueError:
        return False
    else:
        return float(n).is_integer()

# Create function to return a date in datetime format from a string
# Would be used in Function submit_data_costs


def change_date_in_format_datetime(y,m,d):

    global dt

    dt = datetime.datetime(y,m,d,0,0,0)
    return dt

# Create function which will define list containing type of costs based on the
# nature of costs selected on dropdown menu

clicked_nature_costs = StringVar()
clicked_nature_costs.set('Acquisti Personali')



######### Submit Costs Part

def submit_data_costs():

    global list_nature_of_costs
    conn = sqlite3.connect('/Users/Kappadona/Desktop/bookkeeping_v10.db')
    cur = conn.cursor()

    conn_1 = sqlite3.connect('/Users/Kappadona/Desktop/budget_v4.db')
    cur_1 = conn_1.cursor()

    # Get EUR/CHF currency exchange rate at specific date through API

    date_exchange_rate = date_payment_box.get()
    api_request_specific_date_exchange_rate = requests.get('https://api.ratesapi.io/api/'+date_exchange_rate)
    api_specific_date_exchange_rate = json.loads(api_request_specific_date_exchange_rate.content)
    exchange_currency_specific_date =api_specific_date_exchange_rate['rates']['CHF']
    exchange_currency_date = float(exchange_currency_specific_date)


    # Get Date of payment

    date_payment= date_payment_box.get()

    # Create a list out of a date in string format, then insert elements of list
    # into function named "change_date_in_format_datetime"

    list_date_1 = date_payment.split('-')

    y = int(list_date_1[0])
    m = int(list_date_1[1])
    d = int(list_date_1[2])

    date_costs_to_add = change_date_in_format_datetime(y,m,d)

    # Define type of fees to apply to the charge

    fee = 0
    list_values = list(fees.values())
    if clicked_type_of_payment.get() == 'Prepaid/Credit Card':
        fee = list_values[0]
    elif clicked_type_of_payment.get() == 'Cash con ritiro al bancomat':
        fee = list_values[1]
    else:
        fee = 0

    # costs_in_Euro_box.get() returns a string, the string is then converted to an
    # integer or float by using Function named "is_integer(n)"

    costs_in_Eur= costs_in_Euro_box.get()

    integer_or_float = is_integer(costs_in_Eur)

    try:
        if integer_or_float == True:
            costs_in_Eur = int(costs_in_Eur)
        else:
            costs_in_Eur = float(costs_in_Eur)
    except:
        msg = 'Non dimenticare di inserire l\'importo prima di premere il pulsante "Aggiungi Importo del Costo"'
        messagebox.showwarning(message = msg)


    # Once converted to int or float, we are able to add the fees
    # to charges in € converted into CHF

    try:costs_in_CHF = round(((costs_in_Eur*exchange_currency_date) + ((costs_in_Eur*exchange_currency_date)*fee)),2)
    except:
        r1 = random.randint(0,10)
        r2 = random.randint(0,10)
        msg = 'Vediamo quanto sei bravo/brava nelle moltiplicazioni, quanto fa ' + str(r1) + ' multiplicato per ' + str(r2) + ' ?' 
        messagebox.showwarning(message = msg)

    # Insert into Table

    
    try:cur.execute("INSERT INTO bookkeeping_v10 VALUES (:spesa,:natural_del_costo,:costo_EUR,:costo_CHF,:daily_exchange_rate,:data)",
                {
                    'spesa':type_of_costs,
                    'natural_del_costo':clicked_nature_costs.get(),
                    'costo_EUR':costs_in_Euro_box.get(),
                    'costo_CHF':costs_in_CHF,
                    'daily_exchange_rate':exchange_currency_date,
                    'data':date_payment, 
                    })

    except:
        r1 = random.randint(0,10)
        r2 = random.randint(0,10)
        msg = 'Altra domanda quanto fa ' + str(r1) + ' multiplicato per ' + str(r2) + ' ?'
        messagebox.showwarning(message = msg)

    cur.execute("SELECT *, oid FROM bookkeeping_v10")
    records_c = cur.fetchall()

    print(records_c)

    # Check whether costs are over budget limit within a certain period

    cur_1.execute("SELECT *, oid FROM budget_v4")
    records_budget = cur_1.fetchall()

    for i in range (len(records_budget)):
        # Check first if a budget was set

        if records_budget[i][0] == 'Budget generale':

            # Check then if date of payment is within dates of budget

            # Create a list out of the begining date budget recorded in DB
            # such as for instance 2020-04-12 will become [2020,04,12]

            list_begining_date = records_budget[i][2].split('-')

            # Convert each element of list into an integer

            year_begining = int(list_begining_date[0])
            month_begining = int(list_begining_date[1])
            day_begining = int(list_begining_date[2])

            # Final step convert string format of a begining date budget into datetime format

            begining_date = change_date_in_format_datetime(year_begining,month_begining,day_begining)

            # Create a list out of the ending date budget recorded in DB
            # such as for instance 2020-04-14 will become [2020,04,14]

            list_ending_date = records_budget[i][3].split('-')

            # Convert each element of list into an integer

            year_ending = int(list_ending_date[0])
            month_ending = int(list_ending_date[1])
            day_ending= int(list_ending_date[2])

            # Final step to convert string format of a ending date budget into datetime format

            ending_date = change_date_in_format_datetime(year_ending,month_ending,day_ending)

            # Checking if date of payment is within dates of budget

            if begining_date <=date_costs_to_add and date_costs_to_add <= ending_date:

                # Checking if the individual cost is greater than the budget limit

                    if costs_in_CHF> records_budget[i][4]:
                        msg = 'Attenzione il valore del costo supera il budget generale prefissato'
                        messagebox.showwarning(message = msg)

                    cur.execute("SELECT ROUND(SUM(costo_CHF),2) FROM bookkeeping_v10")
                    records_costs_CHF = cur.fetchall()

                    # Checking if the sum of costs in CHF is greater than the budget limit

                    if records_costs_CHF[0][0] > records_budget[i][4]:
                        msg = 'Attenzione hai superato il budget prefissato'
                        messagebox.showwarning(message = msg)

            # Checking if type of budget corresponds to 'Budget per nature dei costi

            elif records_budget[i][0] == 'Budget per natura dei costi':

                # Create a list out of the begining date budget recorded in DB
                # such as for instance 2020-04-12 will become [2020,04,12]

                list_begining_date = records_budget[i][2].split('-')

                # Convert each element of list into an integer
                year_begining = int(list_begining_date[0])
                month_begining = int(list_begining_date[1])
                day_begining = int(list_begining_date[2])

                # Final step convert string format of a begining date budget into datetime format

                begining_date = change_date_in_format_datetime(year_begining,month_begining,day_begining)

                # Create a list out of the ending date budget recorded in DB
                # such as for instance 2020-04-14 will become [2020,04,14]

                list_ending_date = records_budget[i][3].split('-')

                # Convert each element of list into an integer

                year_ending = int(list_ending_date[0])
                month_ending = int(list_ending_date[1])
                day_ending= int(list_ending_date[2])

                # Final step convert string format of a begining date budget into datetime format

                ending_date = change_date_in_format_datetime(year_ending,month_ending,day_ending)

                # Checking if date of budget correspond to date of payment of costs

                if begining_date <=date_costs_to_add and date_costs_to_add<= ending_date:

                    # Checking if the nature of costs selected corresponds to budget of same nature

                    if clicked_nature_costs.get() == records_budget[i][1]:

                        # Checking if the individual cost is greater than the budget limit

                        if costs_in_CHF> records_budget[i][4]:
                            msg = 'Attenzione il valore dell\'importo supera il budget prefissato per il periodo stabilito'
                            messagebox.showwarning(message = msg)

                            cur.execute("SELECT natural_del_costo,costo_CHF, oid FROM bookkeeping_v10")

                            records_costs_nature = cur.fetchall()

                            # Create dictionary which will be populated with values representing all
                            # the costs associated to the nature of costs

                            costs_nature_total={'Acquisti Personali':0, 'Attività Recreative e Intratenimento':0,
                                                'Altro':0,'Comunicazione e Media':0,
                                                'Costi Energetici':0,'Salute': 0, 'Spese Domestiche':0,
                                                'Spostamenti e Transporti':0,'Vacanze e Viaggi':0}

                            for k in costs_nature_total.keys() :
                                for nature_of_costs in records_costs_nature:
                                    if k == nature_of_costs[0]:
                                        costs_nature_total[k] +=nature_of_costs[1]
                            # Once dictionary populated, check if total costs of one kind of nature cost
                            # is greater than the budget limit

                            for k,v in costs_nature_total.items():
                                if k == records_budget[0][1]:
                                    if costs_nature_total[k] > records_budget[i][4]:
                                        msg = 'Attenzione hai superato il budget prefissato stabilito'
                                        messagebox.showwarning(message = msg)
    

    costs_in_Euro_box.delete(0,END)
    date_payment_box.delete(0,END)

    conn_1.commit()
    conn_1.close()

    conn.commit()
    conn.close()    

    


def create_list_type_of_costs():
    elements_costs = clicked_nature_costs.get()
    global list_costs_elements
    list_costs_elements= []
    for k,v in costs_nature.items():
        if k == elements_costs:
            for i in range (len(v)):
                list_costs_elements.append(costs_nature[k][i])

# Create function which will define the type of costs to select from combobox 

def change_values():
        combobox_type_of_costs['values']=list_costs_elements



# Create function Tkinter Combobox Event Binding for type of costs

def callback_function_type_of_costs(event):
    global type_of_costs
    type_of_costs = combobox_type_of_costs.get()
    return type_of_costs
        



def submit_costs_window():

    global submit_costs_window
    global date_payment_box
    global costs_in_Euro_box
    

    submit_costs_window = Tk()
    submit_costs_window.title('Spesa/Costo da Aggiungere')
    submit_costs_window.geometry('600x350')

    # Create functin which will define the type of costs to select from combobox 

    def change_values():

        try:
            combobox_type_of_costs['values']=list_costs_elements

        except:
            msg = 'Premere pulsante "Aggiungi Natura Costi" prima di selezionare tipo di costo'
            messagebox.showwarning(message = msg)



    # Create function Tkinter Combobox Event Binding for type of costs

    def callback_function_type_of_costs(event):

        global type_of_costs
        type_of_costs = combobox_type_of_costs.get()
        return type_of_costs



    # Create label for how payment is carried out

    type_payment_label = Label(submit_costs_window, text = 'Tipo di pagamento effetuato', anchor = W)
    type_payment_label.grid(row = 0, column = 0, sticky = W)

    # Create dropdown menu to select fees to add to cost in CHF
    
    drop_down_box_fees = OptionMenu(submit_costs_window,clicked_type_of_payment, *fees)
    drop_down_box_fees.grid(row =0 , column =1, sticky = W, pady = 15)


    # Create label for nature of charges next to dropdown menu

    nature_cost_label = Label(submit_costs_window, text = 'Seleziona natura del costo', anchor = W)
    nature_cost_label.grid(row = 1, column =0, sticky = W)

    # Create dropdown menu to select nature of costs
    drop_down_box_nature_costs = OptionMenu(submit_costs_window,clicked_nature_costs,*costs_nature)
    drop_down_box_nature_costs.grid(row =1, column = 1,sticky =W)

    # Create button to add nature of costs selected

    submit_nature_costs_button = Button(submit_costs_window,text ='Aggiungi Natura del Costo', command = create_list_type_of_costs)
    submit_nature_costs_button.grid(row = 2, column =0, sticky = W, pady = 12)

    # Create Label to select type of costs

    type_of_costs_label = Label(submit_costs_window, text = 'Seleziona fra le opzioni il tipo di costo', anchor = W)
    type_of_costs_label.grid(row = 3, column = 0, sticky = W)

    # Create combobox to select type of cost from it

    combobox_type_of_costs = ttk.Combobox(submit_costs_window, width = 15, values = ['Accessori','Altro',
                                             'Calzature','Giocatoli',
                                             'Profumi','Vestiti'],
                                      postcommand = change_values)

    combobox_type_of_costs.grid(row = 3, column = 1,sticky = W, pady = 15)

    combobox_type_of_costs.bind('<<ComboboxSelected>>',callback_function_type_of_costs)


    # Create label for date of payment

    date_payment_label = Label(submit_costs_window, text = 'Inserisci data del pagamento nel format XXXX-MM-GG', anchor = W)
    date_payment_label.grid(row = 4, column = 0, sticky = W)

    # Create text box to insert date of payment

    date_payment_box = Entry(submit_costs_window, width = 9)
    date_payment_box.grid(row= 4, column = 1, sticky = W, pady = 10)

    # Create label next to to text box, where amount of the charge in EUR will be inserted

    costs_in_Euro_label = Label(submit_costs_window, text = 'Inserisci importo del costo in €: ', anchor = W)
    costs_in_Euro_label.grid(row =5 , column = 0, sticky = W)

    # Create text box to insert sum in EUR of the charge

    costs_in_Euro_box = Entry(submit_costs_window, width = 12)
    costs_in_Euro_box.grid(row = 5, column = 1, sticky = W, pady = 10)

    # Create button to submit costs

    submit_data_costs_in_DB = Button(submit_costs_window, text = 'Aggiungi Costi', command = submit_data_costs)
    submit_data_costs_in_DB.grid(row = 6, column = 0, sticky = W)

    # Create label to exit

    exit_submit_cost_window_label = Label(submit_costs_window, text = 'Per chiudere, premere tasto EXIT:',anchor = W)
    exit_submit_cost_window_label.grid(row = 7, column = 0, columnspan = 1, sticky = W, pady = 10)

    # Create button Exit

    exit_submit_cost_window_button = Button(submit_costs_window, text = 'Exit', command = submit_costs_window.destroy)
    exit_submit_cost_window_button.grid(row = 7, column = 1, sticky = W)

    submit_costs_window.mainloop()
    

######### Budget Part


def submit_data_budget():

    conn_1 = sqlite3.connect('/Users/Kappadona/Desktop/budget_v4.db')
    cur_1 = conn_1.cursor()

    # Determine what is going to be the nature of the budget into DB,
    # general budget or one of the several nature of costs

    if clicked_budget.get() == 'Budget generale':
        nature_of_budget = clicked_budget.get()
    else:
        nature_of_budget = nature_budget

    # Convert date begining period budget and budget into integer

    budget_begining_date = budget_begining_date_box.get()
    budget_ending_date = budget_ending_date_box.get()

    cur_1.execute("INSERT INTO budget_v4 VALUES (:tipo_budget,:natura_budget,:data_inizio,:data_fine,:somma_budget)",
                {
                    'tipo_budget':clicked_budget.get(),
                    'natura_budget':nature_of_budget,
                    'data_inizio':budget_begining_date,
                    'data_fine':budget_ending_date,
                    'somma_budget':budget_entry_box.get(), 
                    })

    

    budget_entry_box.delete(0,END)
    budget_begining_date_box.delete(0,END)
    budget_ending_date_box.delete(0,END)

    cur_1.execute("SELECT *, oid FROM budget_v4")
    records_b = cur_1.fetchall()

    print(records_b)

    conn_1.commit()
    conn_1.close()
    
def delete_budget_data():

    conn_1 = sqlite3.connect('/Users/Kappadona/Desktop/budget_v4.db')
    cur_1 = conn_1.cursor()

    cur_1.execute("DELETE FROM budget_v4")

    conn_1.commit()
    conn_1.close()
    

# Create function Tkinter Combobox Event Binding for nature of budget

def callback_function_budget(event):
    global nature_budget
    nature_budget = combobox_budget_nature_costs.get()
    return nature_budget
    

def define_budget_window():

    global window_budget_window
    global budget_begining_date_box
    global budget_ending_date_box
    global budget_entry_box
    
    define_budget_window = Tk()
    define_budget_window.title('Budget')
    define_budget_window.geometry('600x600')


    # Create function to use eventualy Combobox to define nature of budget 

    def create_budget_combobox():

        global combobox_budget_nature_costs
        budget_type = clicked_budget.get()
        if budget_type == 'Budget per natura dei costi':

            combobox_budget_nature_costs = ttk.Combobox(define_budget_window, width = 24,
                                                    values = ['Acquisti Personali',
                                                              'Attività Recreative e Intratenimento',
                                                              'Altro','Comunicazione e Media',
                                                              'Costi Energetici',
                                                              'Salute',
                                                              'Spese Domestiche',
                                                              'Spostamenti e Transporti',
                                                              'Vacanze e Viaggi'])

            combobox_budget_nature_costs.grid(row = 1, column = 1,sticky = W, pady = 15)

            combobox_budget_nature_costs.bind('<<ComboboxSelected>>',callback_function_budget)


    

    # Create label for drop down menu type of budget

    type_budget_label = Label(define_budget_window, text = 'Scegli il Tipo di Budget da Definire', anchor = W)
    type_budget_label.grid(row = 0, column = 0, sticky = W)

    # Create dropdown menu for type of budget

    drop_down_box_budget = OptionMenu(define_budget_window,clicked_budget, *list_type_of_budget)
    drop_down_box_budget.grid(row =0 , column =1, sticky = W)

    # Create button to add type of budget selected

    submit_type_budget_button = Button(define_budget_window,text ='Aggiungi il Tipo di Budget', command = create_budget_combobox)
    submit_type_budget_button.grid(row = 1, column =0, sticky = W, pady = 10)

    # Create label for sum of budget to insert

    sum_budget_label = Label(define_budget_window, text = 'Inserisci la somma del budget in CHF', anchor = W)
    sum_budget_label.grid(row = 2, column= 0, sticky = W)

    # Create entry box for sum budget

    budget_entry_box = Entry(define_budget_window, width = 9)
    budget_entry_box.grid(row = 2, column = 1, sticky = W, pady = 10)

    # Create label to inform format date

    info_format_date_label = Label(define_budget_window, text = 'Inserire data nel format XXXX-MM-GG', anchor = W)
    info_format_date_label.grid(row= 3, column = 0, sticky = W)
                        
    # Create label for begining date budget

    begining_date_label = Label(define_budget_window, text = 'Inserisci inizio data periodo budget', anchor = W)
    begining_date_label.grid(row = 4, column = 0, sticky = W)

    # Create entry box for begining date budget

    budget_begining_date_box = Entry(define_budget_window, width = 9)
    budget_begining_date_box.grid(row = 4, column = 1, sticky = W, pady = 10)

    # Create label for ending date budget

    ending_date_label = Label(define_budget_window, text = 'Inserisci fine data periodo budget', anchor = W)
    ending_date_label.grid(row = 5, column = 0, sticky = W)

    # Create entry box for ending date budget

    budget_ending_date_box = Entry(define_budget_window, width = 9)
    budget_ending_date_box.grid(row = 5, column = 1, sticky = W, pady = 10)

    # Create button to submit budget

    submit_budget_button = Button(define_budget_window, text = 'Crea Budget', command = submit_data_budget)
    submit_budget_button.grid(row = 6, column = 0, sticky = W, pady = 10)

    # Create button to delete data into DB budget

    delete_budget_button = Button(define_budget_window, text = 'Elimina Tutti i Budget', command = delete_budget_data)
    delete_budget_button.grid(row = 6, column = 1, sticky = W)

    # Create label to exit define_budget_window

    exit_define_budget_window_label = Label(define_budget_window, text = 'Per chiudere, premere tasto EXIT:',anchor = W)
    exit_define_budget_window_label.grid(row = 7, column = 0, sticky = W, pady = 10)

    # Create button to exit define_budget_window

    exit_define_window_budget_button = Button(define_budget_window, text = 'EXIT', command = define_budget_window.destroy)
    exit_define_window_budget_button.grid(row = 7, column = 1, sticky = W)
    


######### Show Costs Part  


def show_costs_window():

    global show_costs_window
    show_costs_window = Tk()
    show_costs_window.title('Lista Costi €/CHF')
    show_costs_window.geometry('800x600')

    conn = sqlite3.connect('/Users/Kappadona/Desktop/bookkeeping_v10.db')
    cur = conn.cursor()

    cur.execute("SELECT *, oid FROM bookkeeping_v10")
    records = cur.fetchall()

    print_charges= ''
    for charges in records:
        print_charges+= 'N°ID: ' + str(charges[6]) + '\tcosto associato al N°ID: ' + charges[0] +'\n'

    print_charges_EUR =''
    for charges_EUR in records:
        print_charges_EUR+= ' Prezzo in €: ' + str(charges_EUR[2]).ljust(5) + '\tequivalente in CHF: ' + '\n'

        
    print_charges_CHF = ''
    for charges_CHF in records:
        print_charges_CHF+=str(charges_CHF[3]) + '.-' +'\n'

    cur.execute("SELECT ROUND(SUM(costo_EUR),2) FROM bookkeeping_v10")


    sum_charges_EUR = cur.fetchall() # return a tuple

    # convert tuple into a list data type
    sum_charges_EUR_l = [item for t in sum_charges_EUR for item in t]

    cur.execute("SELECT ROUND(SUM(costo_CHF),2) FROM bookkeeping_v10")

    sum_charges_CHF = cur.fetchall()
    
    # convert tuple into a list data type
    sum_charges_CHF_l = [item for t in sum_charges_CHF for item in t]


    # Create label related to variable string print_charges

    query_charges_label = Label(show_costs_window, text = print_charges, anchor = W, justify = 'left')
    query_charges_label.grid(row = 0, column = 0, sticky = W, pady = 10)

    # Create label related to variable string print_charges_EUR

    query_charges_EUR = Label(show_costs_window, text = print_charges_EUR, anchor = W, justify = 'left')
    query_charges_EUR.grid(row = 0, column = 1, sticky = W, pady = 10)

    # Create label related to variable string print_charge_CHF

    query_charges_CHF = Label(show_costs_window, text= print_charges_CHF, anchor = W, justify = 'left')
    query_charges_CHF.grid(row = 0, column = 2, sticky = W)

    # Create label to insert sum_charges_EUR:

    query_total_charges_EUR = Label(show_costs_window, text = 'Costi Totali in €: ' + str(sum_charges_EUR_l[0]), anchor = W, justify = 'left')
    query_total_charges_EUR.grid(row = 1, column = 0, sticky = W)

    # Create label for Costi Totali in CHF:

    query_total_charges_CHF = Label(show_costs_window, text = 'Costi Totali in CHF: ' + str(sum_charges_CHF_l[0])+'.-' , anchor = W, justify = 'left')
    query_total_charges_CHF.grid(row = 2, column = 0, sticky = W)

    # Create label to Exit

    exit_label = Label(show_costs_window, text = 'Per chiudere, premere tasto EXIT:',anchor = W)
    exit_label.grid(row = 3, column = 0, sticky = W, pady = 10)

    # Create button EXIT

    exit_show_costs_window_button = Button(show_costs_window, text = 'Exit', command = show_costs_window.destroy)
    exit_show_costs_window_button.grid(row = 3, column = 1, sticky = W)

    conn.commit()
    conn.close()

    
def delete_costs_window():

    global delete_costs_window

    delete_costs_window = Tk()
    delete_costs_window.title('Azzera Costi')
    delete_costs_window.geometry('325x150')

    def delete_individual_costs():

        conn = sqlite3.connect('/Users/Kappadona/Desktop/bookkeeping_v10.db')
        cur = conn.cursor()

        try:cur.execute("DELETE FROM bookkeeping_v10 WHERE oid= " + delete_individual_costs_ID_box.get())

        except:
            msg = 'Non dimenticare di inserire N°ID prima di premere il pulsante per rimuovere un unico importo'
            messagebox.showwarning(message = msg)

            # cancella_costi.get() enables to retrieve data from data base,
            # the n° ID entered in text box 'cancella_cost',
            # next to label "Inserisci N° ID Spesa da rimuovere

        delete_individual_costs_ID_box.delete(0,END)
        conn.commit()
        conn.close()

    # Delete all data in costs DB

    def delete_all_costs():

        conn = sqlite3.connect('/Users/Kappadona/Desktop/bookkeeping_v10.db')
        cur = conn.cursor()

        # delete all from DB

        cur.execute("DELETE FROM bookkeeping_v10")

        conn.commit()

        conn.close()

    # Create label to delete one by one charge in Eur & CHF next to text box

    delete_individual_costs_ID_label = Label(delete_costs_window, text = 'Inserisci N° ID Spesa da rimuovere: ', pady = 10, anchor = W)
    delete_individual_costs_ID_label.grid(row = 0, column = 0, sticky = W)

    # Create text box to insert n°ID of charge to remove individually

    delete_individual_costs_ID_box = Entry(delete_costs_window, width =4)
    delete_individual_costs_ID_box.grid(row = 0, column = 1, sticky = W)

    # Create button to remove one costs at the time

    delete_individual_costs_ID_button = Button(delete_costs_window, text = 'Rimuovi Singolo Costo', anchor = W,command = delete_individual_costs)
    delete_individual_costs_ID_button.grid(row = 1, column = 0, sticky = W)

    # Create button to remove all costs at once

    delete_all_costs_button= Button(delete_costs_window, text = 'Azzera Tutti i Costi', anchor = W, command = delete_all_costs)
    delete_all_costs_button.grid(row = 2, column = 0, sticky=W, pady = 10)


    # Create label to Exit

    exit_show_costs_window_label = Label(delete_costs_window, text = 'Per chiudere, premere tasto EXIT:',anchor = W)
    exit_show_costs_window_label.grid(row = 3, column = 0, sticky = W, pady = 10)

    # Create button EXIT

    exit_show_costs_window_button = Button(delete_costs_window, text = 'Exit', command = delete_costs_window.destroy)
    exit_show_costs_window_button.grid(row = 3, column = 1, sticky = W)
    

def generate_docs_window():

    global generate_docs_window
    generate_docs_window = Tk()
    generate_docs_window.title('Documenti')
    generate_docs_window.geometry('300x115')


    # Create function to generate excel documents

    def generate_excel_doc():

        conn = sqlite3.connect('/Users/Kappadona/Desktop/bookkeeping_v10.db')
        cur = conn.cursor()

        cur.execute("SELECT *, oid FROM bookkeeping_v10")
        records_for_excel_doc = cur.fetchall()

        os.chdir('/Users/Kappadona/Desktop')

        wb = openpyxl.Workbook()
        sheet = wb.active

        sheet['A2'] = 'Tipo di Costo'
        sheet['A2'].font = Font(bold = True)
        sheet['B1'] = 'Natural del Costo'
        sheet['B1'].font = Font(bold = True)
        sheet['C1'] ='Costo in Euro'
        sheet['C1'].font = Font(bold =True)
        sheet['D1'] = 'Equivalente in CHF'
        sheet['D1'].font = Font(bold = True)
        sheet['E1'] = 'Tasso cambio giornaliero'
        sheet['E1'].font = Font(bold =True)
        sheet['F1'] = 'Data'
        sheet['F1'].font = Font(bold = True)
        sheet['G1'] = 'N°ID'
        sheet['G1'].font = Font(bold = True)

        wb.save('Track_Costi.xlsx')

        # each tuple represent a row
        total_number_of_rows = len(records_for_excel_doc)

        wb = openpyxl.load_workbook('Track_Costi.xlsx')
        sheet = wb.active

        # loop through row, row represent a tuple

        for num_row in range(3,3+total_number_of_rows):
            # nested loop which represent column, There are 4 column in total
            # which represent respectively spesa, costo in Euro, equivalent in CHF
            # and N° ID.

            for column_number in range(1,8):

                # convert tuple into list
                records_for_excel_doc_list = list(records_for_excel_doc[num_row-3])
                # get column letter
                column_letter = get_column_letter(column_number)
                # assign value of i element of list converted from tuple to cell
                sheet[column_letter + str(num_row)] = records_for_excel_doc_list[column_number-1]
        wb.save('Track_Costi.xlsx')

        conn.commit()
        conn.close()

    # Create button to generate an excel doc

    generate_excel_doc_button = Button(generate_docs_window, text = 'Genera Documento Excel', anchor = W, command = generate_excel_doc)
    generate_excel_doc_button.grid(row = 0, column = 0, sticky = W, pady = 10)

    # Create label to Exit

    exit_generate_docs_window_label = Label(generate_docs_window, text = 'Per chiudere, premere tasto EXIT:',anchor = W)
    exit_generate_docs_window_label.grid(row = 1, column = 0, sticky = W, pady = 10)

    # Create button EXIT

    exit_generate_docs_window_button = Button(generate_docs_window, text = 'Exit', command = generate_docs_window.destroy)
    exit_generate_docs_window_button.grid(row = 1, column = 1, sticky = W)

def analysis_costs_window():
    
    global analysis_costs_window
    analysis_costs_window  = Tk()
    analysis_costs_window.title('Analisi Costi')
    analysis_costs_window.geometry('300x125')

    # Function to generate a pie chart

    def costs_nature_pie_chart():

        conn = sqlite3.connect('/Users/Kappadona/Desktop/bookkeeping_v10.db')
        cur = conn.cursor()

        cur.execute("SELECT natural_del_costo,costo_EUR, oid FROM bookkeeping_v10")
        records_costs_nature = cur.fetchall()

        costs_nature_total={'Acquisti Personali':0, 'Attività Recreative e Intratenimento':0,
                            'Altro':0,'Comunicazione e Media':0,
                            'Costi Energetici':0,'Salute': 0, 'Spese Domestiche':0,
                            'Spostamenti e Transporti':0,'Vacanze e Viaggi':0}

        for k in costs_nature_total.keys() :
            for nature_of_costs in records_costs_nature:
                if k == nature_of_costs[0]:
                    costs_nature_total[k] +=nature_of_costs[1]

        list_nature_of_costs = [float(v) for v in costs_nature_total.values()]

        sizes = list_nature_of_costs

        labels = 'Acquisti Personali', 'Attività Recreative e Intratenimento.','Altro','Comunicazione e Media','Costi Energetici','Salute','Spese Domestiche','Spostamenti e Transporti','Vacanze e Viaggi'

        explode = (0,0,0,0,0,0,0,0,0)

        fig1, ax1 = plt.subplots()
        ax1.pie(sizes, explode = explode, labels = labels, autopct = '%1.1f%%',textprops={'fontsize': 10},shadow = True, startangle = 90)
        ax1.axis('equal')
        plt.show()

    def costs_nature_bar_chart():

        conn = sqlite3.connect('/Users/Kappadona/Desktop/bookkeeping_v10.db')
        cur = conn.cursor()

        cur.execute("SELECT natural_del_costo,costo_EUR, oid FROM bookkeeping_v10")
        records_costs_nature = cur.fetchall()

        costs_nature_total={'Acquisti Personali':0, 'Attività Recreative e Intratenimento':0,
                            'Altro':0,'Comunicazione e Media':0,
                            'Costi Energetici':0,'Salute': 0, 'Spese Domestiche':0,
                            'Spostamenti e Transporti':0,'Vacanze e Viaggi':0}

        for k in costs_nature_total.keys() :
            for nature_of_costs in records_costs_nature:
                if k == nature_of_costs[0]:
                    costs_nature_total[k] +=nature_of_costs[1]
        list_nature_of_costs = [float(v) for v in costs_nature_total.values()]

        objects = ('Acquisti Pers.', 'Attività Recre. e Intratenimento.','Altro','Comunicazione e Media','Costi Energetici','Salute','Spese Domestiche','Spostamenti e Transporti','Vacanze e Viaggi')
        y_pos = np.arange(len(objects))
        performance = list_nature_of_costs

        plt.bar(y_pos,performance, align = 'center', alpha = 0.5)
        plt.xticks(y_pos, objects, fontsize = 7)
        plt.ylabel('Costi in € per natura dei costi')
        plt.title('Grafica Natura Costi')
        plt.show()


    # Create button for pie chart
    pie_chart_button=Button(analysis_costs_window, text = 'Pie Chart', anchor = W, command = costs_nature_pie_chart)
    pie_chart_button.grid(row = 0, column = 0, sticky = W, pady = 10)

    # Create button for bar chart

    bar_chart_button = Button(analysis_costs_window, text = 'Bart Chart', anchor = W, command = costs_nature_bar_chart)
    bar_chart_button.grid(row = 1, column = 0, sticky = W, pady = 10)

    # Create label to Exit

    exit_analysis_costs_window_label = Label(analysis_costs_window, text = 'Per chiudere, premere tasto EXIT:',anchor = W)
    exit_analysis_costs_window_label.grid(row = 2, column = 0, sticky = W, pady = 10)

    # Create button EXIT

    exit_analysis_costs_window_button = Button(analysis_costs_window, text = 'Exit', command = analysis_costs_window.destroy)
    exit_analysis_costs_window_button.grid(row = 2, column = 1, sticky = W)
    

# Create welcom Label

welcome_label = Label(root, text = 'Benvenuto, a secondo delle tue esigenze, premi su uno dei seguenti pulsanti', font = font_style)
welcome_label.grid(row = 0, column = 0, columnspan = 2, pady =10)

# Create Button to submit payment

submit_payment_button = Button(root, text = 'Aggiungere Spesa/Costo', anchor = W, command = submit_costs_window)
submit_payment_button.grid(row = 1, column = 0, sticky = W, pady = 10)

# Create Button to define budget

define_budget_button = Button(root, text = 'Definisci Budget', anchor = W, command = define_budget_window)
define_budget_button.grid(row = 1, column = 1, sticky = W, pady = 10, padx =10)

# Create Button to show payment carried out/costs

show_costs_button = Button(root, text = 'Mostra Costi', anchor = W, command = show_costs_window)
show_costs_button.grid(row = 2, column = 0, sticky = W, pady = 10)

# Create Button to delete costs

delete_costs_button = Button(root, text = 'Rimuovere Costi', anchor = W, command = delete_costs_window)
delete_costs_button.grid(row = 2, column = 1, sticky = W, pady = 10, padx =10)

# Create Button to generate Docs

generate_docs_button = Button(root, text = 'Genera Documento', anchor = W, command = generate_docs_window)
generate_docs_button.grid(row = 3, column = 0, sticky = W, pady = 10)

# Create Button to make analysis

analysis_button = Button(root, text = 'Analisi costi', anchor = W, command = analysis_costs_window)
analysis_button.grid(row = 3, column = 1, sticky = W,pady = 10, padx =10)

# Create label to Exit

exit_root_label = Label(root ,text = 'Per chiudere, premere tasto EXIT:',anchor = W)
exit_root_label.grid(row = 4, column = 0, sticky = W, pady = 10)

# Create button EXIT

exit_root_button = Button(root, text = 'Exit', command = root.destroy)
exit_root_button.grid(row = 4, column = 1, sticky = W, padx = 10)


root.mainloop()


